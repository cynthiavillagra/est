"""
Genera el archivo Excel ParkControl_Data.xlsx con las 5 hojas
del sistema, datos de ejemplo y fórmulas relevantes.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference
from datetime import datetime, timedelta
import random

# ============================================
# ESTILOS
# ============================================

BLUE_DARK = "0A0E1A"
BLUE_MED = "1A2236"
BLUE_HEADER = "1E3A5F"
BLUE_ACCENT = "3B82F6"
GREEN = "22C55E"
ORANGE = "F59E0B"
RED = "EF4444"
WHITE = "F0F4FF"
GRAY = "94A3B8"
DARK_BG = "111827"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE_HEADER, end_color=BLUE_HEADER, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

data_font = Font(name="Calibri", size=10, color="333333")
data_align = Alignment(horizontal="left", vertical="center")
center_align = Alignment(horizontal="center", vertical="center")

thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

alt_fill_1 = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

formula_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
formula_font = Font(name="Calibri", size=10, color="92400E", italic=True)

section_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
section_font = Font(name="Calibri", bold=True, color="1E40AF", size=10)


def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def style_data_row(ws, row, num_cols, is_alt=False):
    fill = alt_fill_1 if is_alt else alt_fill_2
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = fill
        cell.border = thin_border


def auto_width(ws, extra=3):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 30)


# ============================================
# CREAR WORKBOOK
# ============================================

wb = openpyxl.Workbook()

# ============================================
# HOJA 1: Configuracion
# ============================================

ws_config = wb.active
ws_config.title = "Configuracion"
ws_config.sheet_properties.tabColor = BLUE_ACCENT

headers = ["campo", "valor", "descripcion"]
ws_config.append(headers)
style_header_row(ws_config, 1, 3)

config_data = [
    ["capacidad_total", 30, "Cantidad total de lugares de estacionamiento"],
    ["tarifa_hora", 500, "Tarifa por hora en pesos"],
    ["tarifa_fraccion", 250, "Tarifa por fracción (30 min) en pesos"],
    ["tarifa_dia", 3500, "Tarifa máxima por día en pesos"],
    ["nombre_parking", "ParkControl Central", "Nombre del estacionamiento"],
    ["direccion", "Av. Corrientes 1234, CABA", "Dirección física"],
    ["telefono", "011-4567-8901", "Teléfono de contacto"],
    ["horario_apertura", "06:00", "Hora de apertura"],
    ["horario_cierre", "23:00", "Hora de cierre"],
    ["moneda", "ARS", "Moneda de operación"],
]

for i, row in enumerate(config_data):
    ws_config.append(row)
    style_data_row(ws_config, i + 2, 3, i % 2 == 0)

auto_width(ws_config)

# ============================================
# HOJA 2: Vehiculos
# ============================================

ws_vehiculos = wb.create_sheet("Vehiculos")
ws_vehiculos.sheet_properties.tabColor = GREEN

headers = ["patente", "titular", "tipo", "habilitado", "abonado", "deuda", "telefono", "email", "fecha_registro"]
ws_vehiculos.append(headers)
style_header_row(ws_vehiculos, 1, len(headers))

vehiculos_data = [
    ["AA123BB", "Juan Pérez", "Auto", "SI", "SI", 0, "11-2345-6789", "jperez@email.com", "2026-01-15"],
    ["AC456DF", "María García", "Auto", "SI", "NO", 3500, "11-3456-7890", "mgarcia@email.com", "2026-02-10"],
    ["AD789GH", "Carlos López", "Camioneta", "SI", "SI", 0, "11-4567-8901", "clopez@email.com", "2026-01-20"],
    ["AE101IJ", "Ana Rodríguez", "Moto", "SI", "NO", 0, "11-5678-9012", "arodriguez@email.com", "2026-03-05"],
    ["AF202KL", "Pablo Díaz", "Auto", "SI", "SI", 0, "11-6789-0123", "pdiaz@email.com", "2026-01-25"],
    ["AD111EF", "Roberto Sánchez", "Auto", "SI", "NO", 2000, "11-7890-1234", "rsanchez@email.com", "2026-02-15"],
    ["AG222HH", "Laura Fernández", "Auto", "SI", "NO", 1500, "11-8901-2345", "lfernandez@email.com", "2026-02-28"],
    ["AI333JJ", "Diego Morales", "Auto", "SI", "NO", 1000, "11-9012-3456", "dmorales@email.com", "2026-03-10"],
    ["AK444LL", "Sofía Martínez", "Auto", "NO", "NO", 5000, "11-0123-4567", "smartinez@email.com", "2026-01-30"],
    ["AM555NN", "Lucas Romero", "Auto", "SI", "SI", 0, "11-1234-5678", "lromero@email.com", "2026-03-15"],
]

for i, row in enumerate(vehiculos_data):
    ws_vehiculos.append(row)
    style_data_row(ws_vehiculos, i + 2, len(headers), i % 2 == 0)

# Validaciones
dv_tipo = DataValidation(type="list", formula1='"Auto,Camioneta,Moto,Bicicleta,Otro"', allow_blank=False)
dv_tipo.error = "Seleccioná un tipo válido"
dv_tipo.prompt = "Tipo de vehículo"
ws_vehiculos.add_data_validation(dv_tipo)
dv_tipo.add(f"C2:C1000")

dv_sn = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
dv_sn.error = "Solo SI o NO"
ws_vehiculos.add_data_validation(dv_sn)
dv_sn.add(f"D2:D1000")
dv_sn.add(f"E2:E1000")

auto_width(ws_vehiculos)

# ============================================
# HOJA 3: Movimientos
# ============================================

ws_mov = wb.create_sheet("Movimientos")
ws_mov.sheet_properties.tabColor = ORANGE

headers = ["patente", "fecha", "hora", "tipo", "monto", "operador", "observaciones"]
ws_mov.append(headers)
style_header_row(ws_mov, 1, len(headers))

# Generar datos de ejemplo para mayo 2026
base_date = datetime(2026, 5, 1)
patentes = ["AA123BB", "AC456DF", "AD789GH", "AE101IJ", "AF202KL",
            "AD111EF", "AG222HH", "AI333JJ", "AM555NN"]

movimientos = []

for day_offset in range(8):  # 1 al 8 de mayo
    date = base_date + timedelta(days=day_offset)
    fecha_str = date.strftime("%Y-%m-%d")

    # Generar entre 8-15 entradas por día
    num_entries = random.randint(8, 15)
    used_patentes = random.sample(patentes, min(num_entries, len(patentes)))

    for pat in used_patentes:
        # Entrada
        hora_e = f"{random.randint(7, 18):02d}:{random.randint(0, 59):02d}:{random.randint(0, 59):02d}"
        movimientos.append([pat, fecha_str, hora_e, "Entrada", 0, "Sistema", ""])

        # Salida (70% de probabilidad)
        if random.random() < 0.7:
            hora_s_h = int(hora_e[:2]) + random.randint(1, 5)
            if hora_s_h > 23:
                hora_s_h = 23
            hora_s = f"{hora_s_h:02d}:{random.randint(0, 59):02d}:{random.randint(0, 59):02d}"
            monto = random.choice([500, 750, 1000, 1250, 1500, 2000])
            movimientos.append([pat, fecha_str, hora_s, "Salida", monto, "Sistema", ""])

# Ordenar por fecha y hora
movimientos.sort(key=lambda x: (x[1], x[2]))

for i, row in enumerate(movimientos):
    ws_mov.append(row)
    style_data_row(ws_mov, i + 2, len(headers), i % 2 == 0)

# Validación para tipo
dv_tipo_mov = DataValidation(type="list", formula1='"Entrada,Salida"', allow_blank=False)
dv_tipo_mov.error = "Solo Entrada o Salida"
ws_mov.add_data_validation(dv_tipo_mov)
dv_tipo_mov.add(f"D2:D5000")

total_mov = len(movimientos)
auto_width(ws_mov)

# ============================================
# HOJA 4: Estadisticas
# ============================================

ws_stats = wb.create_sheet("Estadisticas")
ws_stats.sheet_properties.tabColor = RED

headers = [
    "fecha", "ingresos_dia", "salidas_dia", "recaudacion_dia",
    "ocupacion_max", "hora_pico_inicio", "hora_pico_fin",
    "ingresos_mes_acum", "promedio_diario"
]
ws_stats.append(headers)
style_header_row(ws_stats, 1, len(headers))

# Fórmulas para cada día (del 1 al 8 de mayo)
for day_offset in range(8):
    date = base_date + timedelta(days=day_offset)
    fecha_str = date.strftime("%Y-%m-%d")
    row_num = day_offset + 2

    ws_stats.cell(row=row_num, column=1, value=fecha_str)

    # B: Ingresos del día = COUNTIFS en Movimientos
    ws_stats.cell(row=row_num, column=2,
                  value=f'=COUNTIFS(Movimientos!B:B,A{row_num},Movimientos!D:D,"Entrada")')

    # C: Salidas del día
    ws_stats.cell(row=row_num, column=3,
                  value=f'=COUNTIFS(Movimientos!B:B,A{row_num},Movimientos!D:D,"Salida")')

    # D: Recaudación del día = SUMIFS monto
    ws_stats.cell(row=row_num, column=4,
                  value=f'=SUMIFS(Movimientos!E:E,Movimientos!B:B,A{row_num})')

    # E: Ocupación máxima (estimada: ingresos - salidas previas acumuladas)
    ws_stats.cell(row=row_num, column=5,
                  value=f'=MAX(B{row_num}-C{row_num}+5,B{row_num})')

    # F y G: Hora pico (valores fijos por ahora, se pueden calcular con macros)
    ws_stats.cell(row=row_num, column=6, value="17:00")
    ws_stats.cell(row=row_num, column=7, value="19:00")

    # H: Ingresos acumulados del mes
    ws_stats.cell(row=row_num, column=8, value=f'=SUM(B$2:B{row_num})')

    # I: Promedio diario
    ws_stats.cell(row=row_num, column=9, value=f'=ROUND(H{row_num}/{day_offset + 1},0)')

    style_data_row(ws_stats, row_num, len(headers), day_offset % 2 == 0)

    # Marcar celdas con fórmula
    for col in [2, 3, 4, 5, 8, 9]:
        cell = ws_stats.cell(row=row_num, column=col)
        cell.fill = formula_fill
        cell.font = formula_font

# Fila de totales
total_row = 11
ws_stats.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True, size=11, color="1E40AF")
ws_stats.cell(row=total_row, column=1).fill = section_fill

for col in range(1, len(headers) + 1):
    ws_stats.cell(row=total_row, column=col).fill = section_fill
    ws_stats.cell(row=total_row, column=col).border = thin_border
    ws_stats.cell(row=total_row, column=col).font = Font(bold=True, color="1E40AF")
    ws_stats.cell(row=total_row, column=col).alignment = center_align

ws_stats.cell(row=total_row, column=2, value=f'=SUM(B2:B9)')
ws_stats.cell(row=total_row, column=3, value=f'=SUM(C2:C9)')
ws_stats.cell(row=total_row, column=4, value=f'=SUM(D2:D9)')
ws_stats.cell(row=total_row, column=5, value=f'=MAX(E2:E9)')
ws_stats.cell(row=total_row, column=8, value=f'=H9')
ws_stats.cell(row=total_row, column=9, value=f'=ROUND(AVERAGE(I2:I9),0)')

# Agregar sección de resumen
ws_stats.cell(row=13, column=1, value="📊 RESUMEN DE FÓRMULAS").font = section_font
ws_stats.cell(row=13, column=1).fill = section_fill
ws_stats.merge_cells("A13:I13")

formulas_info = [
    ["ingresos_dia", '=COUNTIFS(Movimientos!B:B, fecha, Movimientos!D:D, "Entrada")', "Cuenta entradas del día"],
    ["salidas_dia", '=COUNTIFS(Movimientos!B:B, fecha, Movimientos!D:D, "Salida")', "Cuenta salidas del día"],
    ["recaudacion_dia", '=SUMIFS(Movimientos!E:E, Movimientos!B:B, fecha)', "Suma montos cobrados ese día"],
    ["ingresos_mes_acum", '=SUM(B$2:B[fila])', "Acumulado de ingresos en el mes"],
    ["promedio_diario", '=ROUND(acumulado / día, 0)', "Promedio de ingresos por día"],
]

ws_stats.cell(row=14, column=1, value="Campo").font = header_font
ws_stats.cell(row=14, column=1).fill = header_fill
ws_stats.cell(row=14, column=2, value="Fórmula (lógica)").font = header_font
ws_stats.cell(row=14, column=2).fill = header_fill
ws_stats.cell(row=14, column=3, value="").fill = header_fill
ws_stats.cell(row=14, column=4, value="Descripción").font = header_font
ws_stats.cell(row=14, column=4).fill = header_fill

for idx, (campo, formula, desc) in enumerate(formulas_info):
    r = 15 + idx
    ws_stats.cell(row=r, column=1, value=campo).font = data_font
    ws_stats.cell(row=r, column=2, value=formula).font = Font(name="Consolas", size=9, color="92400E")
    ws_stats.cell(row=r, column=4, value=desc).font = data_font
    for c in range(1, 5):
        ws_stats.cell(row=r, column=c).fill = alt_fill_1 if idx % 2 == 0 else alt_fill_2
        ws_stats.cell(row=r, column=c).border = thin_border

auto_width(ws_stats, extra=4)

# ============================================
# HOJA 5: Usuarios
# ============================================

ws_users = wb.create_sheet("Usuarios")
ws_users.sheet_properties.tabColor = "A855F7"  # Purple

headers = ["usuario", "password", "nombre", "rol", "activo", "ultimo_login"]
ws_users.append(headers)
style_header_row(ws_users, 1, len(headers))

users_data = [
    ["admin", "admin123", "Administrador", "admin", "SI", "2026-05-08 10:30:00"],
    ["operador", "oper123", "Juan Operador", "operador", "SI", "2026-05-08 08:00:00"],
    ["visor", "visor123", "María Consulta", "visor", "SI", "2026-05-07 14:15:00"],
    ["carlos", "carlos123", "Carlos López", "operador", "SI", "2026-05-06 09:00:00"],
    ["ana", "ana123", "Ana Rodríguez", "visor", "NO", "2026-04-20 11:00:00"],
]

for i, row in enumerate(users_data):
    ws_users.append(row)
    style_data_row(ws_users, i + 2, len(headers), i % 2 == 0)

# Color coding for roles
role_colors = {
    "admin": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
    "operador": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
    "visor": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
}
for r in range(2, len(users_data) + 2):
    rol_val = ws_users.cell(row=r, column=4).value
    if rol_val in role_colors:
        ws_users.cell(row=r, column=4).fill = role_colors[rol_val]

# Inactive user highlighting
for r in range(2, len(users_data) + 2):
    if ws_users.cell(row=r, column=5).value == "NO":
        for c in range(1, len(headers) + 1):
            ws_users.cell(row=r, column=c).font = Font(name="Calibri", size=10, color="9CA3AF", strikethrough=True)

# Validaciones
dv_rol = DataValidation(type="list", formula1='"admin,operador,visor"', allow_blank=False)
dv_rol.error = "Roles válidos: admin, operador, visor"
ws_users.add_data_validation(dv_rol)
dv_rol.add(f"D2:D100")

dv_activo = DataValidation(type="list", formula1='"SI,NO"', allow_blank=False)
ws_users.add_data_validation(dv_activo)
dv_activo.add(f"E2:E100")

# Sección de permisos
ws_users.cell(row=9, column=1, value="🔐 PERMISOS POR ROL").font = section_font
ws_users.cell(row=9, column=1).fill = section_fill
ws_users.merge_cells("A9:F9")

permisos_headers = ["Permiso", "admin", "operador", "visor"]
for i, h in enumerate(permisos_headers):
    ws_users.cell(row=10, column=i + 1, value=h).font = header_font
    ws_users.cell(row=10, column=i + 1).fill = header_fill
    ws_users.cell(row=10, column=i + 1).alignment = header_align
    ws_users.cell(row=10, column=i + 1).border = thin_border

permisos = [
    ["Ver dashboard", "✅", "✅", "✅"],
    ["Ver estadísticas", "✅", "✅", "✅"],
    ["Abrir barrera", "✅", "✅", "❌"],
    ["Ver historial", "✅", "✅", "❌"],
    ["Ver vehículos", "✅", "✅", "❌"],
    ["Gestionar vehículos", "✅", "❌", "❌"],
    ["Ver deudas", "✅", "✅", "❌"],
    ["Cobrar deudas", "✅", "❌", "❌"],
    ["Ver reportes", "✅", "❌", "❌"],
    ["Configuración", "✅", "❌", "❌"],
    ["Gestionar usuarios", "✅", "❌", "❌"],
]

for idx, perm in enumerate(permisos):
    r = 11 + idx
    for c, val in enumerate(perm):
        cell = ws_users.cell(row=r, column=c + 1, value=val)
        cell.font = data_font
        cell.alignment = center_align
        cell.fill = alt_fill_1 if idx % 2 == 0 else alt_fill_2
        cell.border = thin_border

auto_width(ws_users)

# ============================================
# HOJA DASHBOARD (resumen con fórmulas)
# ============================================

ws_dash = wb.create_sheet("Dashboard_Resumen")
ws_dash.sheet_properties.tabColor = "14B8A6"  # Teal

# Título
ws_dash.merge_cells("A1:F1")
ws_dash.cell(row=1, column=1, value="📊 PARKCONTROL - RESUMEN EN TIEMPO REAL").font = Font(
    name="Calibri", bold=True, size=14, color="1E40AF"
)
ws_dash.cell(row=1, column=1).fill = section_fill
ws_dash.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

# KPIs
kpis = [
    ["Capacidad Total", '=VLOOKUP("capacidad_total",Configuracion!A:B,2,FALSE)'],
    ["Vehículos Registrados", '=COUNTA(Vehiculos!A:A)-1'],
    ["Vehículos Habilitados", '=COUNTIF(Vehiculos!D:D,"SI")'],
    ["Vehículos con Deuda", '=COUNTIF(Vehiculos!F2:F1000,">"&0)'],
    ["Deuda Total", '=SUMIF(Vehiculos!F2:F1000,">"&0,Vehiculos!F2:F1000)'],
    ["Total Movimientos", '=COUNTA(Movimientos!A:A)-1'],
    ["Total Entradas", '=COUNTIF(Movimientos!D:D,"Entrada")'],
    ["Total Salidas", '=COUNTIF(Movimientos!D:D,"Salida")'],
    ["Recaudación Total", '=SUM(Movimientos!E:E)'],
    ["Usuarios Activos", '=COUNTIF(Usuarios!E:E,"SI")'],
    ["Abonados", '=COUNTIF(Vehiculos!E:E,"SI")'],
    ["Tarifa por Hora", '=VLOOKUP("tarifa_hora",Configuracion!A:B,2,FALSE)'],
]

ws_dash.cell(row=3, column=1, value="INDICADOR").font = header_font
ws_dash.cell(row=3, column=1).fill = header_fill
ws_dash.cell(row=3, column=1).border = thin_border
ws_dash.cell(row=3, column=2, value="VALOR").font = header_font
ws_dash.cell(row=3, column=2).fill = header_fill
ws_dash.cell(row=3, column=2).border = thin_border
ws_dash.cell(row=3, column=3, value="FÓRMULA APLICADA").font = header_font
ws_dash.cell(row=3, column=3).fill = header_fill
ws_dash.cell(row=3, column=3).border = thin_border

for idx, (label, formula) in enumerate(kpis):
    r = 4 + idx
    ws_dash.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
    ws_dash.cell(row=r, column=2, value=formula)
    ws_dash.cell(row=r, column=2).fill = formula_fill
    ws_dash.cell(row=r, column=2).font = formula_font
    ws_dash.cell(row=r, column=3, value=formula.replace("=", "")).font = Font(
        name="Consolas", size=8, color="6B7280"
    )
    for c in range(1, 4):
        ws_dash.cell(row=r, column=c).fill = alt_fill_1 if idx % 2 == 0 else alt_fill_2
        ws_dash.cell(row=r, column=c).border = thin_border

# Top 5 deudores con fórmula
r_start = 18
ws_dash.merge_cells(f"A{r_start}:F{r_start}")
ws_dash.cell(row=r_start, column=1, value="💰 TOP DEUDORES").font = section_font
ws_dash.cell(row=r_start, column=1).fill = section_fill

for i, h in enumerate(["#", "Patente", "Titular", "Deuda"]):
    ws_dash.cell(row=r_start + 1, column=i + 1, value=h).font = header_font
    ws_dash.cell(row=r_start + 1, column=i + 1).fill = header_fill
    ws_dash.cell(row=r_start + 1, column=i + 1).border = thin_border

# Deudores ordenados manualmente (LARGE no funciona bien para lookup cruzado en Excel sin arrays avanzados)
ws_dash.cell(row=r_start + 2, column=1, value=1)
ws_dash.cell(row=r_start + 2, column=2, value="AK444LL")
ws_dash.cell(row=r_start + 2, column=3, value='=VLOOKUP(B20,Vehiculos!A:B,2,FALSE)')
ws_dash.cell(row=r_start + 2, column=4, value='=VLOOKUP(B20,Vehiculos!A:F,6,FALSE)')

ws_dash.cell(row=r_start + 3, column=1, value=2)
ws_dash.cell(row=r_start + 3, column=2, value="AC456DF")
ws_dash.cell(row=r_start + 3, column=3, value='=VLOOKUP(B21,Vehiculos!A:B,2,FALSE)')
ws_dash.cell(row=r_start + 3, column=4, value='=VLOOKUP(B21,Vehiculos!A:F,6,FALSE)')

ws_dash.cell(row=r_start + 4, column=1, value=3)
ws_dash.cell(row=r_start + 4, column=2, value="AD111EF")
ws_dash.cell(row=r_start + 4, column=3, value='=VLOOKUP(B22,Vehiculos!A:B,2,FALSE)')
ws_dash.cell(row=r_start + 4, column=4, value='=VLOOKUP(B22,Vehiculos!A:F,6,FALSE)')

ws_dash.cell(row=r_start + 5, column=1, value=4)
ws_dash.cell(row=r_start + 5, column=2, value="AG222HH")
ws_dash.cell(row=r_start + 5, column=3, value='=VLOOKUP(B23,Vehiculos!A:B,2,FALSE)')
ws_dash.cell(row=r_start + 5, column=4, value='=VLOOKUP(B23,Vehiculos!A:F,6,FALSE)')

ws_dash.cell(row=r_start + 5 + 1, column=1, value=5)
ws_dash.cell(row=r_start + 5 + 1, column=2, value="AI333JJ")
ws_dash.cell(row=r_start + 5 + 1, column=3, value='=VLOOKUP(B24,Vehiculos!A:B,2,FALSE)')
ws_dash.cell(row=r_start + 5 + 1, column=4, value='=VLOOKUP(B24,Vehiculos!A:F,6,FALSE)')

for r in range(r_start + 2, r_start + 7):
    for c in range(1, 5):
        ws_dash.cell(row=r, column=c).border = thin_border
        ws_dash.cell(row=r, column=c).alignment = center_align
    ws_dash.cell(row=r, column=4).fill = formula_fill

auto_width(ws_dash, extra=5)
ws_dash.column_dimensions["C"].width = 45

# ============================================
# GUARDAR
# ============================================

output_path = "ParkControl_Data.xlsx"
wb.save(output_path)
print(f"✅ Archivo creado: {output_path}")
print(f"   📊 Hojas: Configuracion, Vehiculos, Movimientos ({total_mov} registros), Estadisticas, Usuarios, Dashboard_Resumen")
print(f"   📐 Fórmulas: COUNTIFS, SUMIFS, VLOOKUP, SUM, COUNTA, COUNTIF, AVERAGE, MAX, ROUND")
print(f"   ✔️  Validaciones: listas desplegables para tipo, habilitado, abonado, rol")
